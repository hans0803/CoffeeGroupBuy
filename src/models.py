#!/usr/bin/env python3
"""
資料庫模型
使用 SQLite 儲存產品與訂單資料
"""

import sqlite3
import json
import os
from datetime import datetime, timezone, timedelta
from typing import Optional, List, Dict, Tuple
from contextlib import contextmanager
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dotenv import load_dotenv

load_dotenv()

DATABASE_URL = os.environ.get("DATABASE_URL", "").strip() or None
IS_POSTGRES = bool(DATABASE_URL)

if IS_POSTGRES:
    import psycopg2
    import psycopg2.extras


# UTC+8 時區
TZ_UTC8 = timezone(timedelta(hours=8))

# 資料庫在專案根目錄的 data 資料夾
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATABASE_PATH = os.path.join(BASE_DIR, 'data', 'products.db')


def get_now_utc8() -> datetime:
    """取得 UTC+8 當前時間"""
    return datetime.now(TZ_UTC8)


def ensure_data_dir():
    """確保 data 目錄存在"""
    data_dir = os.path.dirname(DATABASE_PATH)
    os.makedirs(data_dir, exist_ok=True)


@contextmanager
def get_db():
    """取得資料庫連線的 context manager (支援 SQLite 或 PostgreSQL)"""
    if IS_POSTGRES:
        conn = psycopg2.connect(DATABASE_URL)
        # Ensure cursors use RealDictCursor for dict-like row access
        conn.cursor_factory = psycopg2.extras.RealDictCursor
        try:
            yield conn
        finally:
            conn.close()
    else:
        ensure_data_dir()
        conn = sqlite3.connect(DATABASE_URL_LOCAL if 'DATABASE_URL_LOCAL' in globals() else DATABASE_PATH)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()

def row_to_dict(cursor, row):
    """將資料庫 row 轉換為 dict，相容多種驅動與 cursor 類型"""
    if row is None: return None
    if isinstance(row, dict): return dict(row)
    try:
        return dict(row)
    except (TypeError, ValueError):
        # Fallback for generic tuples if cursor description is available
        if cursor and hasattr(cursor, 'description'):
            return {desc[0]: val for desc, val in zip(cursor.description, row)}
    return row

def db_execute(cursor, query: str, params: tuple = (), fetch: str = None):
    """
    通用 SQL 執行函式：
    自動將 SQLite 的 '?' 轉換為 PostgreSQL 的 '%s'
    """
    if IS_POSTGRES:
        query = query.replace('?', '%s')
    
    cursor.execute(query, params)
    
    if fetch == 'all':
        return cursor.fetchall()
    elif fetch == 'one':
        return cursor.fetchone()
    return None


def init_db():
    """初始化資料庫表格"""
    with get_db() as conn:
        if IS_POSTGRES:
            cursor = conn.cursor()
        else:
            cursor = conn.cursor()

        # 產品表 - 注意：PostgreSQL 沒有 TIMESTAMP DEFAULT CURRENT_TIMESTAMP 的寫法，需要改寫
        auto_inc = "SERIAL" if IS_POSTGRES else "INTEGER AUTOINCREMENT"
        
        db_execute(cursor, f'''
            CREATE TABLE IF NOT EXISTS products (
                id TEXT PRIMARY KEY,
                sku TEXT,
                name TEXT NOT NULL,
                price INTEGER NOT NULL,
                original_price INTEGER,
                image_url TEXT,
                product_url TEXT,
                category TEXT NOT NULL,
                category_name TEXT,
                roast TEXT,
                processing TEXT,
                purchase_count INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 訂單表
        db_execute(cursor, f'''
            CREATE TABLE IF NOT EXISTS orders (
                id {auto_inc} PRIMARY KEY,
                customer_name TEXT NOT NULL,
                items_json TEXT NOT NULL,
                total INTEGER NOT NULL,
                submitted_to_google BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 評論資料表
        db_execute(cursor, f'''
            CREATE TABLE IF NOT EXISTS reviews (
                id {auto_inc} PRIMARY KEY,
                product_id TEXT NOT NULL,
                reviewer_name TEXT NOT NULL,
                rating INTEGER NOT NULL CHECK (rating >= 1 AND rating <= 5),
                comment TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
            )
        ''')
        
        # 簡單遷移：針對 SQLite 新增欄位 (PostgreSQL 初次建立就會包含，若要完善需處理 PG 遷移)
        if not IS_POSTGRES:
            cursor.execute("PRAGMA table_info(products)")
            columns = [info[1] for info in cursor.fetchall()]

            if 'roast' not in columns:
                cursor.execute("ALTER TABLE products ADD COLUMN roast TEXT")

            if 'purchase_count' not in columns:
                cursor.execute("ALTER TABLE products ADD COLUMN purchase_count INTEGER DEFAULT 0")
        
        conn.commit()
        print("資料庫初始化完成")
        print("資料庫初始化完成")


def save_products(products: List[Dict]):
    """儲存產品到資料庫（更新或插入，並保留商品建立時間）"""
    with get_db() as conn:
        cursor = conn.cursor()

        for p in products:
            if IS_POSTGRES:
                query = '''
                    INSERT INTO products
                    (id, sku, name, price, original_price, image_url, product_url, category, category_name, roast, processing, updated_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT(id) DO UPDATE SET
                        sku=EXCLUDED.sku,
                        name=EXCLUDED.name,
                        price=EXCLUDED.price,
                        original_price=EXCLUDED.original_price,
                        image_url=EXCLUDED.image_url,
                        product_url=EXCLUDED.product_url,
                        category=EXCLUDED.category,
                        category_name=EXCLUDED.category_name,
                        roast=EXCLUDED.roast,
                        processing=EXCLUDED.processing,
                        updated_at=EXCLUDED.updated_at
                '''
            else:
                query = '''
                    INSERT INTO products
                    (id, sku, name, price, original_price, image_url, product_url, category, category_name, roast, processing, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(id) DO UPDATE SET
                        sku=excluded.sku,
                        name=excluded.name,
                        price=excluded.price,
                        original_price=excluded.original_price,
                        image_url=excluded.image_url,
                        product_url=excluded.product_url,
                        category=excluded.category,
                        category_name=excluded.category_name,
                        roast=excluded.roast,
                        processing=excluded.processing,
                        updated_at=excluded.updated_at
                '''
            cursor.execute(query, (
                p['id'], p['sku'], p['name'], p['price'], p.get('original_price'),
                p['image_url'], p['product_url'], p['category'], p['category_name'],
                p.get('roast'), p.get('processing'),
                datetime.now()
            ))

        conn.commit()
        print(f"已儲存 {len(products)} 個產品到資料庫")


# 烘焙度分組定義
ROAST_GROUPS = {
    'light': {
        'label': '淺中',
        'values': ['黃金烘焙', '黃金烘培', '淺烘焙', '淺烘培', '淺焙', '中淺焙', '中淺烘焙']
    },
    'medium': {
        'label': '中焙',
        'values': ['白金烘焙', '白金烘培', '中烘焙', '中烘培', '中焙']
    },
    'medium_dark': {
        'label': '中深',
        'values': ['中深烘焙', '中深烘培', '中深焙']
    },
    'dark': {
        'label': '深焙',
        'values': ['黑金烘焙', '黑金烘培', '深烘焙', '深烘培', '深焙']
    }
}

def get_all_products(category: Optional[str] = None, min_price: Optional[int] = None, max_price: Optional[int] = None,
                    roast: Optional[str] = None, processing: Optional[str] = None) -> List[Dict]:
    """取得所有產品 (支援分類、價格、烘焙度、處理法篩選)"""
    with get_db() as conn:
        cursor = conn.cursor()

        query = "SELECT * FROM products"
        params = []
        conditions = []

        if category:
            conditions.append("category = ?")
            params.append(category)

        if min_price is not None:
            conditions.append("price >= ?")
            params.append(min_price)

        if max_price is not None:
            conditions.append("price <= ?")
            params.append(max_price)

        if roast:
            # 檢查是否為群組篩選
            if roast in ROAST_GROUPS:
                placeholders = ','.join(['?'] * len(ROAST_GROUPS[roast]['values']))
                conditions.append(f"roast IN ({placeholders})")
                params.extend(ROAST_GROUPS[roast]['values'])
            else:
                conditions.append("roast = ?")
                params.append(roast)

        if processing:
            conditions.append("processing = ?")
            params.append(processing)

        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        if category:
            query += " ORDER BY purchase_count DESC, price ASC, name ASC"
        else:
            query += " ORDER BY category, purchase_count DESC, price ASC, name ASC"

        rows = db_execute(cursor, query, tuple(params), fetch='all')
        if IS_POSTGRES:
            return [{k: v for k, v in zip([desc[0] for desc in cursor.description], row)} for row in rows]
        return [dict(row) for row in rows]

def get_newest_products(days: int = 7) -> List[Dict]:
    """取得最近 N 天內首次建立的產品 (即本次新品)"""
    with get_db() as conn:
        cursor = conn.cursor()
        
        if IS_POSTGRES:
            query = """
                SELECT * FROM products
                WHERE created_at >= NOW() - INTERVAL '%s days'
                ORDER BY purchase_count DESC, created_at DESC, category, name ASC
            """
            rows = db_execute(cursor, query, (days,), fetch='all')
        else:
            query = """
                SELECT * FROM products
                WHERE created_at >= datetime('now', ?)
                ORDER BY purchase_count DESC, created_at DESC, category, name ASC
            """
            rows = db_execute(cursor, query, (f'-{days} days',), fetch='all')
        
        # 如果最近 N 天內沒有任何新品，為了避免頁面空白，
        # 我們退而求其次：找出「最新資料庫中 created_at 最大的那一批」
        if not rows:
            if IS_POSTGRES:
                fallback_query = '''
                    SELECT * FROM products
                    WHERE DATE(created_at) = (
                        SELECT DATE(MAX(created_at)) FROM products
                    )
                    ORDER BY category, name ASC
                '''
            else:
                fallback_query = '''
                    SELECT * FROM products
                    WHERE date(created_at) = (
                        SELECT date(max(created_at)) FROM products
                    )
                    ORDER BY category, name ASC
                '''
            rows = db_execute(cursor, fallback_query, fetch='all')
            
        if IS_POSTGRES:
            return [{k: v for k, v in zip([desc[0] for desc in cursor.description], row)} for row in rows]
        return [dict(row) for row in rows]

def update_sales_statistics():
    """統計訂單，更新產品購買次數，並產生銷售統計數據 (統計 JSON)"""
    with get_db() as conn:
        cursor = conn.cursor()

        # 1. 統計所有訂單的商品數量
        products_counts = defaultdict(int)

        cursor.execute("SELECT items_json FROM orders")
        orders = cursor.fetchall()

        for (items_json,) in orders:
            try:
                items = json.loads(items_json)
                for item in items:
                    name = item.get('name')
                    qty = item.get('quantity', 0)
                    if name and qty > 0:
                        products_counts[name] += qty
            except (json.JSONDecodeError, TypeError):
                continue
        
        # 移除先前的更新資料庫(累加)邏輯，因為現在改為即時更新 purchase_count
        conn.commit()

        # 3. 產生統計數據 (從最新的 DB 狀態查詢)
        stats = {
            "top_products": [],
            "roast_stats": {},
            "price_stats": {}
        }

        # A. Top Sellers (Filter by 'drip' as requested)
        cursor.execute('''
            SELECT name, price, image_url, category, purchase_count, roast
            FROM products
            WHERE purchase_count > 0 AND category = 'drip'
            ORDER BY purchase_count DESC
            LIMIT 10
        ''')
        stats["top_products"] = [dict(row) for row in cursor.fetchall()]

        # B. Roast Stats (based on purchase_count, filtered by drip)
        cursor.execute("SELECT roast, purchase_count FROM products WHERE purchase_count > 0 AND category = 'drip'")
        roast_rows = cursor.fetchall()

        # 簡單映射烘焙度 (Align with ROAST_GROUPS)
        roast_counts = defaultdict(int)

        for r_val, count in roast_rows:
            label = "其他"
            if not r_val:
                label = "未分類"
            else:
                if '黃金' in r_val or '淺' in r_val:
                    label = "淺中"  # Light
                elif '白金' in r_val or '中' in r_val and '深' not in r_val:
                    label = "中焙"  # Medium
                elif '中深' in r_val:
                    label = "中深"  # Medium-Dark
                elif '深' in r_val and '中' not in r_val:
                    label = "深焙"  # Dark

            roast_counts[label] += count

        stats["roast_stats"] = dict(roast_counts)

        # C. Price Stats (Exact Price Points, filtered by drip)
        cursor.execute("SELECT price, purchase_count FROM products WHERE purchase_count > 0 AND category = 'drip'")
        price_rows = cursor.fetchall()

        # 使用明確價格統計
        price_exact = defaultdict(int)

        for price, count in price_rows:
            price_exact[str(price)] += count

        # Sort by Price (Key) numerically
        sorted_prices = sorted(price_exact.items(), key=lambda x: int(x[0]))
        stats["price_stats"] = dict(sorted_prices)

        json_path = os.path.join(BASE_DIR, 'data', 'statistics.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(stats, f, ensure_ascii=False, indent=2)

        print(f"已更新銷售統計，寫入 {json_path}")


def get_common_prices(category: Optional[str] = None, limit: int = 12) -> List[int]:
    """取得該分類下常見的價格 (用於快速篩選)"""
    with get_db() as conn:
        cursor = conn.cursor()

        query = """
            SELECT price, COUNT(*) as count
            FROM products
            WHERE 1=1
        """
        params = []

        if category:
            query += " AND category = ?"
            params.append(category)

        # 只取有意義的價格
        query += " GROUP BY price ORDER BY price ASC LIMIT ?"
        params.append(limit)

        rows = db_execute(cursor, query, tuple(params), fetch='all')
        return [row_to_dict(cursor, row)['price'] for row in rows]


def get_current_sales_statistics() -> dict:
    """從目前訂單(orders表)即時計算銷售統計 (不讀取 json)"""
    # 取得資料庫中所有未匯出的訂單
    orders = get_all_orders()

    # 用於累計
    product_counts = defaultdict(int)
    product_info_map = {} # cache name -> {price, category, roast, image_url}

    # 預先取得所有產品詳細資訊以供查詢 (避免 N+1)
    all_products = get_all_products()
    for p in all_products:
        product_info_map[p['name']] = p

    for order in orders:
        items = order['items'] # get_all_orders 已經 parse json
        
        # 增加防禦性機制：確保 items 是列表
        if isinstance(items, dict):
            items = list(items.values())
            
        for item in items:
            if not isinstance(item, dict):
                continue
                
            name = item.get('name')
            qty = item.get('quantity', 0)

            # 找到對應產品資訊 (檢查 category)
            p_info = product_info_map.get(name)
            if p_info and p_info.get('category') == 'drip':
                product_counts[name] += qty

    # 產生統計數據結構
    stats = {
        "top_products": [],
        "roast_stats": {},
        "price_stats": {}
    }

    # 1. Top Products (List)
    temp_top = []
    for name, count in product_counts.items():
        p_info = product_info_map.get(name)
        if p_info:
            temp_top.append({
                "name": name,
                "price": p_info['price'],
                "image_url": p_info['image_url'],
                "category": p_info['category'],
                "purchase_count": count,
                "roast": p_info.get('roast')
            })

    # Sort by count desc
    temp_top.sort(key=lambda x: x['purchase_count'], reverse=True)
    stats["top_products"] = temp_top[:10]

    # 2. Roast Stats
    roast_counts = defaultdict(int)
    for name, count in product_counts.items():
        p_info = product_info_map.get(name)
        r_val = p_info.get('roast') if p_info else None

        label = "其他"
        if not r_val:
            label = "未分類"
        else:
            if '黃金' in r_val or '淺' in r_val:
                label = "淺中"
            elif '白金' in r_val or '中' in r_val and '深' not in r_val:
                label = "中焙"
            elif '中深' in r_val:
                label = "中深"
            elif '深' in r_val and '中' not in r_val:
                label = "深焙"

        roast_counts[label] += count
    stats["roast_stats"] = dict(roast_counts)

    # 3. Price Stats (Exact)
    price_exact = defaultdict(int)
    for name, count in product_counts.items():
        p_info = product_info_map.get(name)
        price = p_info.get('price', 0) if p_info else 0
        if price > 0:
            price_exact[str(price)] += count

    sorted_prices = sorted(price_exact.items(), key=lambda x: int(x[0]))
    stats["price_stats"] = dict(sorted_prices)

    return stats



def get_product_facets(category: str) -> dict:
    """取得該分類下的烘焙度與處理法選項"""
    facets = {'roast': [], 'processing': []}

    with get_db() as conn:
        cursor = conn.cursor()

        # 取得烘焙度
        rows = db_execute(cursor, '''
            SELECT roast, COUNT(*) as count
            FROM products
            WHERE category = ? AND roast IS NOT NULL AND roast != ''
            GROUP BY roast
            ORDER BY count DESC
        ''', (category,), fetch='all')
        facets['roast'] = [row_to_dict(cursor, row)['roast'] for row in rows]

        # 取得處理法
        rows = db_execute(cursor, '''
            SELECT processing, COUNT(*) as count
            FROM products
            WHERE category = ? AND processing IS NOT NULL AND processing != ''
            GROUP BY processing
            ORDER BY count DESC
        ''', (category,), fetch='all')
        facets['processing'] = [row_to_dict(cursor, row)['processing'] for row in rows]

    return facets


def get_product_by_id(product_id: str) -> Optional[dict]:
    """根據 ID 取得單一產品"""
    with get_db() as conn:
        cursor = conn.cursor()
        row = db_execute(cursor, 'SELECT * FROM products WHERE id = ?', (product_id,), fetch='one')
        return row_to_dict(cursor, row)


def get_products_by_ids(product_ids: List[str]) -> List[Dict]:
    """根據多個 ID 取得產品"""
    if not product_ids:
        return []

    with get_db() as conn:
        cursor = conn.cursor()
        placeholders = ','.join('?' * len(product_ids))
        rows = db_execute(cursor, f'SELECT * FROM products WHERE id IN ({placeholders})', tuple(product_ids), fetch='all')
        return [row_to_dict(cursor, row) for row in rows]


def get_categories() -> List[Dict]:
    """取得所有分類及其產品數量"""
    with get_db() as conn:
        cursor = conn.cursor()
        rows = db_execute(cursor, '''
            SELECT category, category_name, COUNT(*) as count
            FROM products
            GROUP BY category, category_name
            ORDER BY category
        ''', fetch='all')
        return [row_to_dict(cursor, row) for row in rows]


def create_order(customer_name: str, items: List[Dict], total: int) -> int:
    """建立新訂單"""
    with get_db() as conn:
        cursor = conn.cursor()
        now = get_now_utc8().strftime('%Y-%m-%d %H:%M:%S')
        
        items_json = json.dumps(items, ensure_ascii=False)
        
        if IS_POSTGRES:
            query = '''
                INSERT INTO orders (customer_name, items_json, total, created_at)
                VALUES (%s, %s, %s, %s) RETURNING id
            '''
            # In RealDictCursor, fetchone() returns a dict e.g. {'id': 1}
            cursor.execute(query, (customer_name, items_json, total, now))
            row = cursor.fetchone()
            order_id = row['id'] if isinstance(row, dict) else row[0]
        else:
            query = '''
                INSERT INTO orders (customer_name, items_json, total, created_at)
                VALUES (?, ?, ?, ?)
            '''
            cursor.execute(query, (customer_name, items_json, total, now))
            order_id = cursor.lastrowid
        
        # 即時更新商品的已售出數量
        for item in items:
            name = item.get('name')
            qty = item.get('quantity', 0)
            if name and qty > 0:
                if IS_POSTGRES:
                    db_execute(cursor, 'UPDATE products SET purchase_count = COALESCE(purchase_count, 0) + %s WHERE name = %s', (qty, name))
                else:
                    db_execute(cursor, 'UPDATE products SET purchase_count = IFNULL(purchase_count, 0) + ? WHERE name = ?', (qty, name))
        
        conn.commit()
        return order_id


def get_all_orders() -> List[Dict]:
    """取得所有訂單"""
    with get_db() as conn:
        cursor = conn.cursor()
        rows = db_execute(cursor, 'SELECT * FROM orders ORDER BY created_at DESC', fetch='all')

        orders = []
        for row in rows:
            order = row_to_dict(cursor, row)
            try:
                items_data = json.loads(order['items_json'])
                # 強制轉換為列表格式，確保相容性
                if isinstance(items_data, dict):
                    order['items'] = list(items_data.values())
                else:
                    order['items'] = items_data
            except (json.JSONDecodeError, TypeError):
                order['items'] = []
            orders.append(order)

        return orders


def mark_order_submitted(order_id: int):
    """標記訂單已提交到 Google"""
    with get_db() as conn:
        cursor = conn.cursor()
        val = True if IS_POSTGRES else 1
        db_execute(cursor, 'UPDATE orders SET submitted_to_google = ? WHERE id = ?', (val, order_id))
        conn.commit()


def get_order_statistics() -> dict:
    """取得訂單統計"""
    with get_db() as conn:
        cursor = conn.cursor()

        # 總訂單數
        row = db_execute(cursor, 'SELECT COUNT(*) as count FROM orders', fetch='one')
        total_orders = row_to_dict(cursor, row)['count'] if row else 0

        # 總金額
        row = db_execute(cursor, 'SELECT SUM(total) as total_amount FROM orders', fetch='one')
        res = row_to_dict(cursor, row)
        total_amount = (res['total_amount'] if res else 0) or 0

        # 各使用者統計
        rows = db_execute(cursor, '''
            SELECT customer_name, COUNT(*) as order_count, SUM(total) as total_spent
            FROM orders
            GROUP BY customer_name
            ORDER BY total_spent DESC
        ''', fetch='all')
        user_stats = [row_to_dict(cursor, row) for row in rows]

        return {
            'total_orders': total_orders,
            'total_amount': total_amount,
            'user_stats': user_stats
        }


def export_orders_to_xlsx(output_path: str = None) -> str:
    """匯出訂單到 Excel 檔案"""
    if output_path is None:
        # 預設輸出到 data 目錄
        ensure_data_dir()
        timestamp = get_now_utc8().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(os.path.dirname(DATABASE_PATH), f'orders_{timestamp}.xlsx')

    orders = get_all_orders()

    # 建立工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "訂單明細"

    # 樣式設定
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 標題列
    headers = ["訂單編號", "顧客姓名", "商品名稱", "單價", "數量", "小計", "訂單總額", "建立時間"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 寫入資料
    row = 2
    for order in orders:
        items = order.get('items', [])
        order_total = order.get('total', 0)
        created_at = order.get('created_at', '')

        for idx, item in enumerate(items):
            subtotal = item.get('price', 0) * item.get('quantity', 0)

            ws.cell(row=row, column=1, value=order['id']).border = thin_border
            ws.cell(row=row, column=2, value=order['customer_name']).border = thin_border
            ws.cell(row=row, column=3, value=item.get('name', '')).border = thin_border
            ws.cell(row=row, column=4, value=item.get('price', 0)).border = thin_border
            ws.cell(row=row, column=5, value=item.get('quantity', 0)).border = thin_border
            ws.cell(row=row, column=6, value=subtotal).border = thin_border

            # 只在第一個商品顯示訂單總額和時間
            if idx == 0:
                ws.cell(row=row, column=7, value=order_total).border = thin_border
                ws.cell(row=row, column=8, value=created_at).border = thin_border
            else:
                ws.cell(row=row, column=7, value="").border = thin_border
                ws.cell(row=row, column=8, value="").border = thin_border

            row += 1

    # 調整欄寬
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20

    # 儲存
    wb.save(output_path)
    print(f"訂單已匯出至: {output_path}")
    return output_path


def clear_orders() -> int:
    """清空所有訂單，回傳刪除數量"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as count FROM orders')
        row = cursor.fetchone()
        count = row['count'] if row else 0
        cursor.execute('DELETE FROM orders')
        conn.commit()
        print(f"已清空 {count} 筆訂單")
        return count


def export_orders() -> Tuple[str, str]:
    """
    匯出訂單到兩個 Excel 檔案

    Returns:
        tuple: (商家訂購單路徑, 內部分發單路徑)
    """
    ensure_data_dir()
    timestamp = get_now_utc8().strftime('%Y%m%d_%H%M%S')

    orders = get_all_orders()

    # 樣式設定
    header_font = Font(bold=True, color="FFFFFF")
    header_fill_blue = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_fill_green = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ========================================
    # 報表一：商家訂購單 (產品統計)
    # ========================================
    vendor_path = os.path.join(os.path.dirname(DATABASE_PATH), f'order_vendor_{timestamp}.xlsx')

    # 統計各產品總數量
    product_summary = defaultdict(lambda: {'quantity': 0, 'price': 0})
    for order in orders:
        for item in order.get('items', []):
            name = item.get('name', '')
            qty = item.get('quantity', 0)
            price = item.get('price', 0)
            product_summary[name]['quantity'] += qty
            product_summary[name]['price'] = price  # 記錄單價

    wb_vendor = Workbook()
    ws_vendor = wb_vendor.active
    ws_vendor.title = "商家訂購單"

    # 標題列
    vendor_headers = ["商品名稱", "數量", "單價", "小計"]
    for col, header in enumerate(vendor_headers, 1):
        cell = ws_vendor.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_green
        cell.alignment = header_alignment
        cell.border = thin_border

    # 寫入資料
    row = 2
    grand_total = 0
    total_qty = 0
    for name, data in sorted(product_summary.items()):
        subtotal = data['price'] * data['quantity']
        grand_total += subtotal
        total_qty += data['quantity']

        ws_vendor.cell(row=row, column=1, value=name).border = thin_border
        ws_vendor.cell(row=row, column=2, value=data['quantity']).border = thin_border
        ws_vendor.cell(row=row, column=3, value=data['price']).border = thin_border
        ws_vendor.cell(row=row, column=4, value=subtotal).border = thin_border
        row += 1

    # 合計列
    ws_vendor.cell(row=row, column=1, value="合計").font = Font(bold=True)
    ws_vendor.cell(row=row, column=1).border = thin_border
    ws_vendor.cell(row=row, column=2, value=total_qty).font = Font(bold=True)
    ws_vendor.cell(row=row, column=2).border = thin_border
    ws_vendor.cell(row=row, column=3, value="").border = thin_border
    ws_vendor.cell(row=row, column=4, value=grand_total).font = Font(bold=True)
    ws_vendor.cell(row=row, column=4).border = thin_border

    # 調整欄寬
    ws_vendor.column_dimensions['A'].width = 50
    ws_vendor.column_dimensions['B'].width = 10
    ws_vendor.column_dimensions['C'].width = 10
    ws_vendor.column_dimensions['D'].width = 12

    wb_vendor.save(vendor_path)
    print(f"商家訂購單已匯出至: {vendor_path}")

    # ========================================
    # 報表二：內部分發單 (完整訂單明細 - 按人名分組)
    # ========================================
    internal_path = os.path.join(os.path.dirname(DATABASE_PATH), f'order_internal_{timestamp}.xlsx')

    # 按姓名分組訂單
    grouped_orders = group_orders_by_customer(orders)

    wb_internal = Workbook()
    ws_internal = wb_internal.active
    ws_internal.title = "內部分發單"

    # 標題列
    internal_headers = ["姓名", "商品名稱", "數量", "單價", "小計", "個人總額"]
    for col, header in enumerate(internal_headers, 1):
        cell = ws_internal.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_blue
        cell.alignment = header_alignment
        cell.border = thin_border

    # 寫入資料
    row = 2
    for customer, data in sorted(grouped_orders.items()):
        customer_items = data['items']
        customer_total = data['total']

        # 合併同名商品
        merged_items = defaultdict(lambda: {'quantity': 0, 'price': 0})
        for item in customer_items:
            name = item.get('name', '')
            qty = item.get('quantity', 0)
            price = item.get('price', 0)
            merged_items[name]['quantity'] += qty
            merged_items[name]['price'] = price

        sorted_items = sorted(merged_items.items())

        for idx, (name, item_data) in enumerate(sorted_items):
            qty = item_data['quantity']
            price = item_data['price']
            subtotal = qty * price

            # 第一欄：姓名 (只顯示一次)
            if idx == 0:
                cell = ws_internal.cell(row=row, column=1, value=customer)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                # 只有第一行顯示個人總額
                ws_internal.cell(row=row, column=6, value=customer_total).border = thin_border
            else:
                ws_internal.cell(row=row, column=1, value="").border = thin_border
                ws_internal.cell(row=row, column=6, value="").border = thin_border

            ws_internal.cell(row=row, column=2, value=name).border = thin_border
            ws_internal.cell(row=row, column=3, value=qty).border = thin_border
            ws_internal.cell(row=row, column=4, value=price).border = thin_border
            ws_internal.cell(row=row, column=5, value=subtotal).border = thin_border

            row += 1

        # 每個人的訂單之間加分隔線或是空行 (選擇加粗下框線區隔)
        for col in range(1, 7):
            cell = ws_internal.cell(row=row-1, column=col)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='medium')
            )

    # 調整欄寬
    ws_internal.column_dimensions['A'].width = 15
    ws_internal.column_dimensions['B'].width = 45
    ws_internal.column_dimensions['C'].width = 8
    ws_internal.column_dimensions['D'].width = 10
    ws_internal.column_dimensions['E'].width = 10
    ws_internal.column_dimensions['F'].width = 12

    wb_internal.save(internal_path)
    print(f"內部分發單已匯出至: {internal_path}")

    return vendor_path, internal_path


def group_orders_by_customer(orders: List[Dict]) -> dict:
    """
    將訂單按客戶姓名分組
    """
    grouped = defaultdict(lambda: {'items': [], 'total': 0})

    for order in orders:
        name = order.get('customer_name', '').strip()
        if not name:
            continue

        items = order.get('items', [])
        total = order.get('total', 0)

        grouped[name]['items'].extend(items)
        grouped[name]['total'] += total

    return grouped


def get_orders_by_customer(name: str) -> List[Dict]:
    """根據姓名取得訂單"""
    with get_db() as conn:
        cursor = conn.cursor()
        rows = db_execute(cursor, 
            'SELECT * FROM orders WHERE customer_name = ? ORDER BY created_at DESC',
            (name,), fetch='all'
        )

        orders = []
        for row in rows:
            order = row_to_dict(cursor, row)
            order['items'] = json.loads(order['items_json'])
            orders.append(order)

        return orders


def get_all_customer_names() -> List[str]:
    """取得所有已下單的客戶姓名"""
    with get_db() as conn:
        cursor = conn.cursor()
        rows = db_execute(cursor, 'SELECT DISTINCT customer_name FROM orders ORDER BY customer_name', fetch='all')
        names = []
        for row in rows:
            d = row_to_dict(cursor, row)
            if isinstance(d, dict) and 'customer_name' in d:
                names.append(d['customer_name'])
            else:
                names.append(row[0]) # Fallback for flat tuples
        return names


def delete_order(order_id: int) -> bool:
    """刪除整筆訂單"""
    with get_db() as conn:
        cursor = conn.cursor()
        
        row = db_execute(cursor, 'SELECT items_json FROM orders WHERE id = ?', (order_id,), fetch='one')
        if row:
            items_json = row['items_json']
            
            items = json.loads(items_json)
            for item in items:
                name = item.get('name')
                qty = item.get('quantity', 0)
                if name and qty > 0:
                    if IS_POSTGRES:
                        db_execute(cursor, 'UPDATE products SET purchase_count = GREATEST(0, COALESCE(purchase_count, 0) - %s) WHERE name = %s', (qty, name))
                    else:
                        db_execute(cursor, 'UPDATE products SET purchase_count = MAX(0, IFNULL(purchase_count, 0) - ?) WHERE name = ?', (qty, name))
                    
        db_execute(cursor, 'DELETE FROM orders WHERE id = ?', (order_id,))
        conn.commit()
        return cursor.rowcount > 0


def update_order_item(order_id: int, product_name: str, new_quantity: int) -> bool:
    """更新訂單項目數量 (如果數量<=0 則刪除該項目)"""
    with get_db() as conn:
        cursor = conn.cursor() # Get cursor from connection

        # 先取得目前訂單內容
        db_execute(cursor, 'SELECT items_json, total FROM orders WHERE id = ?', (order_id,))
        row = cursor.fetchone()

        if not row:
            return False

        items_json = row['items_json']
        current_total = row['total']
        
        items = json.loads(items_json)

        # 尋找並更新項目
        new_items = []
        new_total = 0
        updated = False

        for item in items:
            if item['name'] == product_name:
                old_qty = item.get('quantity', 0)
                delta = new_quantity - old_qty
                
                if new_quantity > 0:
                    item['quantity'] = new_quantity
                    new_items.append(item)
                    new_total += item['price'] * new_quantity
                
                # 如果 quantity <= 0，就略過不加入 (即刪除)
                updated = True
                
                # 即時更新商品的已售出數量
                if delta != 0:
                    if IS_POSTGRES:
                        cursor.execute('UPDATE products SET purchase_count = GREATEST(0, COALESCE(purchase_count, 0) + %s) WHERE name = %s', (delta, product_name))
                    else:
                        cursor.execute('UPDATE products SET purchase_count = MAX(0, IFNULL(purchase_count, 0) + ?) WHERE name = ?', (delta, product_name))
            else:
                new_items.append(item)
                new_total += item['price'] * item['quantity']

        if not updated: # 沒找到該商品，可能是要從購物車邏輯看是否新增？目前邏輯僅支援修改既有
             return False

        # 如果所有商品都被刪光了，是否要刪除整筆訂單？
        # 使用者需求是「個別品項刪除」，如果刪光了應該留著空訂單或是刪除訂單
        # 這裡選擇如果沒商品了，就刪除整筆訂單
        if not new_items:
            db_execute(cursor, 'DELETE FROM orders WHERE id = ?', (order_id,))
        else:
            db_execute(cursor, 
                'UPDATE orders SET items_json = ?, total = ? WHERE id = ?',
                (json.dumps(new_items, ensure_ascii=False), new_total, order_id)
            )

        conn.commit()
        return True


# ==========================================
# 評論相關功能
# ==========================================

def add_review(product_id: str, reviewer_name: str, rating: int, comment: str = None) -> int:
    """新增評論"""
    with get_db() as conn:
        cursor = conn.cursor()
        now = get_now_utc8().strftime('%Y-%m-%d %H:%M:%S')
        
        if IS_POSTGRES:
            query = '''
                INSERT INTO reviews (product_id, reviewer_name, rating, comment, created_at)
                VALUES (%s, %s, %s, %s, %s) RETURNING id
            '''
            cursor.execute(query, (product_id, reviewer_name, rating, comment, now))
            row = cursor.fetchone()
            review_id = row['id']
        else:
            query = '''
                INSERT INTO reviews (product_id, reviewer_name, rating, comment, created_at)
                VALUES (?, ?, ?, ?, ?)
            '''
            cursor.execute(query, (product_id, reviewer_name, rating, comment, now))
            review_id = cursor.lastrowid
            
        conn.commit()
        return review_id


def get_reviews_by_product(product_id: str) -> List[Dict]:
    """取得特定產品的所有評論"""
    with get_db() as conn:
        cursor = conn.cursor()
        
        # 使用 COALESCE 處理可能為 NULL 的 comment 欄位 (確保輸出的 JSON 不會意外變成 null，這邊可以直接用 SQL 處理也可以用 Python 處理)
        # 這裡為了簡單我們先用 Python 處理
        db_execute(cursor, '''
            SELECT id, reviewer_name, rating, comment, created_at
            FROM reviews
            WHERE product_id = ?
            ORDER BY created_at DESC
        ''', (product_id,))
        
        rows = cursor.fetchall()
        reviews = []
        for row in rows:
            r = row_to_dict(cursor, row)
            if r.get('comment') is None:
                r['comment'] = ""
            reviews.append(r)
            
        return reviews


def get_product_review_stats(product_id: str) -> dict:
    """取得單一商品的評論統計(平均星數、評論總數)"""
    with get_db() as conn:
        cursor = conn.cursor()
        if IS_POSTGRES:
            db_execute(cursor, '''
                SELECT AVG(rating)::numeric as avg_rating, COUNT(*) as review_count
                FROM reviews
                WHERE product_id = %s
            ''', (product_id,))
        else:
            db_execute(cursor, '''
                SELECT AVG(rating) as avg_rating, COUNT(*) as review_count
                FROM reviews
                WHERE product_id = ?
            ''', (product_id,))
        row = cursor.fetchone()
        
        avg = row['avg_rating'] if row else 0.0
        count = row['review_count'] if row else 0

        avg = float(avg) if avg is not None else 0.0
        count = int(count) if count is not None else 0
        
        return {
            'avg_rating': round(avg, 1),
            'review_count': count
        }


def get_all_product_review_stats() -> Dict[str, Dict]:
    """
    取得所有產品的評論統計資料 (平均星數與評論數)
    Returns:
        Dict: a mapping from product_id to {'average_rating': float, 'review_count': int}
    """
    with get_db() as conn:
        cursor = conn.cursor()
        
        if IS_POSTGRES:
            cursor.execute('''
                SELECT product_id, 
                       ROUND(AVG(rating)::numeric, 1) as avg_rating, 
                       COUNT(*) as r_count
                FROM reviews
                GROUP BY product_id
            ''')
        else:
            cursor.execute('''
                SELECT product_id, 
                       ROUND(AVG(rating), 1) as avg_rating, 
                       COUNT(*) as r_count
                FROM reviews
                GROUP BY product_id
            ''')
            
        rows = cursor.fetchall()
        
        stats = {}
        for row in rows:
            p_id = row['product_id']
            avg = row['avg_rating']
            cnt = row['r_count']
            
            # 安全轉換，避免 NoneType 錯誤
            stats[p_id] = {
                'average_rating': float(avg) if avg is not None else 0.0,
                'review_count': int(cnt) if cnt is not None else 0
            }
            
        return stats


if __name__ == "__main__":
    init_db()
    print("資料庫初始化完成")
